import pandas as pd
import numpy as np
import io
import logging
from typing import Dict, Tuple, List
from concurrent.futures import ThreadPoolExecutor
import threading
import os
import pytz
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime, timezone
from dateutil.parser import parse

from AzureOperations import AzureOperations
from SharePointOperations import SharePointOperations

logging.basicConfig(level=logging.INFO)

class StockStatusReportOps:
    
    def __init__(self, exported_file_name: str, exported_file_bytes: bytes):
        self.azure_ops = AzureOperations()
        access_token = self.azure_ops.get_access_token()

        self.sharepoint_ops = SharePointOperations(access_token=access_token)

        self.product_list_prefix = os.environ.get('PRODUCT_LIST_PREFIX')
        self.ssr_summary_prefix = os.environ.get('SSR_SUMMARY_PREFIX')

        self.exported_file_name = exported_file_name
        self.exported_file_bytes = exported_file_bytes
        self.stock_status_report_sheet = None
        self.columns_to_hide = [
                                "E",
                                "F",
                                "I",
                                "J",
                                "L",
                                "M",
                                "N",
                                "P",
                                "R",
                                "S",
                                "U",
                                "V",
                                "W",
                                "X",
                                "Y",
                                "Z",
                                "AA",
                                "AB",
                                "AD",
                                "AE",
                                "AF",
                                "AG",
                                "AH",
                                "AI",
                                "AJ",
                                "AK",
                                "AL",
                                "AM",
                                "AN",
                                "AO",
                                "AP",
                                "AQ",
                                "AR",
                                "AS",
                                "AT",
                                "AU",
                                "AV",
                                "AW",
                            ]
        self.new_columns = [
                                "UNIT PRICE",
                                "SOH Value xgst",
                                "PO Cost xgst",
                                "On Order Cost xgst"
                            ]
        self.columns_to_clip = ['qty_onhand', 'qty_SO', 'qty_PO']
        self.australia_now = datetime.now(pytz.timezone('Australia/Sydney'))

        self.ssr_summary_bytes, self.product_list = self.get_ssr_summary_and_product_list()
        self.customer_config = self.get_customer_config(ssr_summary_bytes=self.ssr_summary_bytes)

        self.copy_customer_config = self.customer_config.copy(deep=True)
        self.copy_customer_config.drop(index=0, inplace=True)
        self.copy_customer_config[1] = self.copy_customer_config[1].apply(lambda x: [kw.strip() for kw in x.split(',')]).to_list()

    def start_automate(self) -> Tuple[bytes, str, str, str]:
        logging.info("[StockStatusReportOps] Automating SSR")

        notification_message = ''

        try:
            exported_file_df = self.excel_bytes_as_df()

            logging.info("[StockStatusReportOps] Selecting client rows")
            # Filter rows to select data from clients only
            client_filter = [item for sublist in self.copy_customer_config[1] for item in sublist]
            client_rows_df = exported_file_df[exported_file_df['item_cat1'].astype(str).str.strip().astype(str).str.lower().isin({s.lower() for s in client_filter})]

            logging.info("[StockStatusReportOps] Cleaning up rows with samples")
            # Remove rows that are all samples
            client_rows_wo_samples_df = client_rows_df[~client_rows_df.apply(lambda row: row.astype(str).str.contains('sample', case=False, na=False)).any(axis=1)]

            logging.info("[StockStatusReportOps] Adding new columns i.e. SOH, SO, PO")
            # Add new columns
            client_rows_wo_samples_df.loc[:, self.new_columns] = np.nan
            client_rows_wo_samples_df.loc[:, ["active in web"]] = np.nan.__str__()
            client_rows_wo_samples_df.loc[:, ["CHECK FOR DUPLICATES"]] = np.nan

            logging.info("[StockStatusReportOps] Converting digits negative values in SOH, SO, PO to zero")
            # Set zeroes to negative values
            client_rows_wo_samples_df.loc[:, self.columns_to_clip] = client_rows_wo_samples_df.loc[:, self.columns_to_clip].clip(lower=0)

            logging.info("[StockStatusReportOps] Varifying if any barcodes are missing")
            # Check if any barcodes are missing
            empty_barcodes_df = client_rows_wo_samples_df[client_rows_wo_samples_df['barcode'].isnull()]
            if not empty_barcodes_df.empty:
                logging.info("[StockStatusReportOps] Records with no barcodes found! Building notification message")

                id_list = empty_barcodes_df['ID'].tolist()
                name_list = empty_barcodes_df['Name'].tolist()

                notification_message += "Missing Barcodes for following exported items: -newline-"

                for id_, name_ in zip(id_list, name_list):
                    notification_message += "-tab- - {id_}: {name_} -newline-".format(id_=id_, name_=name_)

                notification_message += "-newline-"

            logging.info("[StockStatusReportOps] Preparing lookup map for unit price and active in web")
            # Build lookup map
            unitPrice_mapping = self.product_list.groupby('barcode')['unitPrice'].first()
            activeInWeb_mapping = self.product_list.groupby('barcode')['ActiveInWeb'].first()

            logging.info("[StockStatusReportOps] Performing lookup of unit price using barcode")
            # Lookup unit price
            client_rows_wo_samples_df.loc[:,'barcode'] = client_rows_wo_samples_df.loc[:,'barcode'].astype(str)
            client_rows_wo_samples_df.loc[:,'UNIT PRICE'] = client_rows_wo_samples_df.loc[:,'barcode'].astype(str).map(unitPrice_mapping).fillna(np.nan)

            logging.info("[StockStatusReportOps] Performing lookup of active in web using barcode")
            # Lookup active in web
            client_rows_wo_samples_df.loc[:,'active in web'] = client_rows_wo_samples_df.loc[:,'active in web'].astype(str)
            client_rows_wo_samples_df.loc[:,'active in web'] = client_rows_wo_samples_df.loc[:,'barcode'].astype(str).map(activeInWeb_mapping).fillna('0')

            # check if no lookup value retrieved for rows with SOH, SO, and PO
            no_lookup_filtered_df = client_rows_wo_samples_df[
                (client_rows_wo_samples_df['qty_onhand'] > 0) &
                (client_rows_wo_samples_df['qty_SO'] > 0) &
                (client_rows_wo_samples_df['qty_PO'] > 0) &
                (client_rows_wo_samples_df['UNIT PRICE'] == np.nan)
            ]

            logging.info("[StockStatusReportOps] Validating rows with no lookup but has SOH, SO, and PO > 1")
            if not no_lookup_filtered_df.empty:
                logging.info("[StockStatusReportOps] Records with no lookup data where SOH, SO, and PO > 1 found! Building notification message")

                id_list = no_lookup_filtered_df['ID'].tolist()
                name_list = no_lookup_filtered_df['Name'].tolist()

                notification_message += "No Lookup Data for following export items where SOH, SO, and PO > 0: -newline-"

                for id_, name_ in zip(id_list, name_list):
                    notification_message += "-tab- - {id_}: {name_} -newline-".format(id_=id_, name_=name_)

                notification_message += "-newline-"

                # Skip rows without looked up price
                logging.info("[StockStatusReportOps] Skipping row without unit price")
                client_rows_wo_samples_df[:, :] = client_rows_wo_samples_df[~client_rows_wo_samples_df.index.isin(no_lookup_filtered_df.index)] 

            logging.info("[StockStatusReportOps] Replacing Null in unit price to zero")
            # Replace NA values with zero, these items has SOH, SO, and PO equals to zero
            client_rows_wo_samples_df.loc[:,'UNIT PRICE'] = client_rows_wo_samples_df.loc[:,'UNIT PRICE'].replace(np.nan, 0)

            logging.info("[StockStatusReportOps] Calculating SOH, PO, and On Order cost")
            # Calculate SOH, PO, and On Order cost
            client_rows_wo_samples_df.loc[:,'SOH Value xgst'] = client_rows_wo_samples_df.loc[:,'UNIT PRICE'] * client_rows_wo_samples_df.loc[:,'qty_onhand']
            client_rows_wo_samples_df.loc[:,'PO Cost xgst'] = client_rows_wo_samples_df.loc[:,'UNIT PRICE'] * client_rows_wo_samples_df.loc[:,'qty_PO']
            client_rows_wo_samples_df.loc[:,'On Order Cost xgst'] = client_rows_wo_samples_df.loc[:,'UNIT PRICE'] * client_rows_wo_samples_df.loc[:,'qty_SO']

            excel_file_bytes, sum_per_client_sheet, ssr_filename = self.build_excel_file_buffer(
                cleaned_ssr_df=client_rows_wo_samples_df
            )

            logging.info("[StockStatusReportOps] Will run read and update SSR Summary function asynchronously")
            thread = threading.Thread(target=self.read_update_ssr_summary, args=(sum_per_client_sheet, ))
            thread.start()

            notification_message = f"Stock Status Report Automation now done! -newline- -newline-" + notification_message

            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

            logging.info("[StockStatusReportOps] SSR Summary automation ends")

            return excel_file_bytes, ssr_filename, mimetype, notification_message

        except Exception as e:
            logging.exception(f"[Stock Status Report Operation] Failure to automate stock status report. Error: {e}")
            return False

    def excel_bytes_as_df(self) -> pd.DataFrame:
        excel_file = io.BytesIO(self.exported_file_bytes)
        return pd.read_excel(excel_file, engine='openpyxl')
    
    def df_to_excel_bytes(self, sheets: Dict[str, pd.DataFrame]) -> bytes:
        output_buffer = io.BytesIO()

        with pd.ExcelWriter(output_buffer, engine='xlsxwriter') as writer:

            for _sheet_name, _sheet_df in sheets.items():

                _sheet_name = _sheet_name[:31]
                cleaned_df = _sheet_df.dropna(how='all').reset_index(drop=True)
                cleaned_df.to_excel(writer, sheet_name=_sheet_name, index=False)

        return output_buffer.getvalue()

    def get_product_list(self) -> pd.DataFrame:
        logging.info("[StockStatusReportOps] Getting product list")

        site_id = self.sharepoint_ops.get_site_id()
        drive_id = self.sharepoint_ops.get_drive_id(site_id=site_id)
        csv_bytes = self.sharepoint_ops.get_bytes_for_latest_file_with_prefix(prefix=self.product_list_prefix, drive_id=drive_id)

        product_list = pd.read_csv(io.BytesIO(csv_bytes))

        return product_list

    def get_ssr_summary_bytes(self) -> bytes:
        logging.info("[StockStatusReportOps] Getting SSR Summary file bytes")

        site_id = self.sharepoint_ops.get_site_id()
        drive_id = self.sharepoint_ops.get_drive_id(site_id=site_id)
        excel_bytes = self.sharepoint_ops.get_bytes_for_latest_file_with_prefix(prefix=self.ssr_summary_prefix, drive_id=drive_id)

        return excel_bytes

    def get_ssr_summary_df(self, ssr_summary_bytes: bytes) -> pd.DataFrame:
        logging.info("[StockStatusReportOps] Getting SSR Summary Dataframe")

        fiscal_year_start, fiscal_year_end = self.get_start_end_fiscal_year()

        main_sheet_title = f"JULY {fiscal_year_start} - JUNE {fiscal_year_end} FY"

        ssr_summary_df = pd.read_excel(io.BytesIO(ssr_summary_bytes), sheet_name=main_sheet_title, header=None, engine='openpyxl')

        return ssr_summary_df

    def get_customer_config(self, ssr_summary_bytes: bytes) -> pd.DataFrame:
        logging.info("[StockStatusReportOps] Getting SSR Customer Config")

        customer_config = pd.read_excel(io.BytesIO(ssr_summary_bytes), sheet_name="customer_config", header=None, engine='openpyxl')

        return customer_config

    def build_excel_file_buffer(self, cleaned_ssr_df: pd.DataFrame) -> Tuple[bytes, Dict, str]:
        logging.info("[StockStatusReportOps] Building excel file buffer from dataframe")

        def process_sheet(sheet_name: str, category, df: pd.DataFrame) -> Tuple[str, pd.DataFrame, Dict]:
            category_set = {str(cat).lower() for cat in category}
            filtered_df = df[df['item_cat1_lower'].isin(category_set)]
            return (
                sheet_name,
                filtered_df,
                {
                    "soh_value_sum": filtered_df['SOH Value xgst'].sum(),
                    "po_cost_sum": filtered_df['PO Cost xgst'].sum(),
                    "so_cost_sum": filtered_df['On Order Cost xgst'].sum()
                }
            )

        sum_per_client_sheet = {}

        # Normalize item_cat1 once for efficiency
        cleaned_ssr_df = cleaned_ssr_df.copy(deep=True)
        cleaned_ssr_df['item_cat1_lower'] = cleaned_ssr_df['item_cat1'].astype(str).str.strip().str.lower()

        # Combine new row and config
        new_row = pd.DataFrame([["CURRENT CUSTOMERS", np.nan, np.nan]])
        config_df = pd.concat([new_row, self.copy_customer_config], ignore_index=True)
        sheet_config = [(row[0], row[1]) for row in config_df.itertuples(index=False, name=None)]

        results: List[Tuple[str, pd.DataFrame, Dict]] = []

        # Threaded filtering and aggregation
        with ThreadPoolExecutor(max_workers=4) as executor:
            futures = []
            for sheet_name, category in sheet_config:
                if str(sheet_name).upper() == 'CURRENT CUSTOMERS':
                    results.append((sheet_name, cleaned_ssr_df, {}))
                else:
                    futures.append(executor.submit(process_sheet, sheet_name, category, cleaned_ssr_df))

            for f in futures:
                results.append(f.result())

        output_buffer = io.BytesIO()

        # Write to Excel (must be done serially)
        with pd.ExcelWriter(output_buffer, engine='xlsxwriter') as writer:  # Consider 'xlsxwriter' for speed
            for sheet_name, df, sums in results:
                df.drop(columns='item_cat1_lower', errors='ignore').to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]

                for col_letter in self.columns_to_hide:
                    try:
                        col_idx = column_index_from_string(col_letter) - 1  # Convert A -> 0, B -> 1, etc.
                        worksheet.set_column(col_idx, col_idx, None, None, {'hidden': True})
                    except Exception as e:
                        logging.warning(f"Could not hide column {col_letter} in {sheet_name}: {e}")

                if sums:
                    sum_per_client_sheet[sheet_name] = sums

        output_buffer.seek(0)

        ssr_filename = f"STOCK STATUS REPORT {self.australia_now.strftime('%Y%m%d')}.xlsx"       

        return output_buffer.getvalue(), sum_per_client_sheet, ssr_filename
    
    def read_update_ssr_summary(self, sum_per_client_sheet: Dict):
        logging.info(f"[StockStatusReportOps] Trying to read and update SSR Summary")

        fiscal_year_start, fiscal_year_end = self.get_start_end_fiscal_year()

        main_sheet_title = f"JULY {fiscal_year_start} - JUNE {fiscal_year_end} FY"

        ssr_summary_df = self.get_ssr_summary_df(ssr_summary_bytes=self.ssr_summary_bytes)

        ssr_summary_wb = load_workbook(io.BytesIO(self.ssr_summary_bytes))

        ssr_summary_sheet = ssr_summary_wb[main_sheet_title]

        new_ssr_summary_dict: Dict[str, Dict] = self.generate_new_ssr_summary(sum_per_client_sheet)

        target_col_idx = self.get_target_ssr_summary_table_column(ssr_summary_df=ssr_summary_df)

        for client_code, sum_dict in new_ssr_summary_dict.items():
            ssr_summary_client_name = self.customer_config[self.customer_config[0] == client_code][2].iloc[0]

            client_row = self.get_target_client_row(ssr_summary_df=ssr_summary_df, ssr_customer_label=ssr_summary_client_name)

            if not isinstance(client_row, int) or not isinstance(target_col_idx, int):
                continue

            for row_string, ssr_sum_fig in sum_dict.items():

                fig_row = self.get_target_client_row(ssr_summary_df=ssr_summary_df, ssr_customer_label=row_string, start_idx=client_row)
                fig_row += 1

                ssr_summary_sheet.cell(row=fig_row, column=target_col_idx + 1).value = ssr_sum_fig

                if isinstance(ssr_sum_fig, (int, float)):
                    ssr_summary_sheet.cell(row=fig_row, column=target_col_idx + 1).number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'


        excel_buffer = io.BytesIO()

        ssr_summary_filename = f"DINA Stock Status Report Overview FY{str(fiscal_year_start)[2:]}-{str(fiscal_year_end)[2:]}.xlsx"

        # Write both DataFrames to different sheets in the same in-memory Excel file
        ssr_summary_wb.save(excel_buffer)

        # Don't forget to rewind the buffer before using it
        excel_buffer.seek(0)

        site_id = self.sharepoint_ops.get_site_id()
        drive_id = self.sharepoint_ops.get_drive_id(site_id=site_id)        
        self.sharepoint_ops.upload_excel_file(drive_id=drive_id, excel_filename=ssr_summary_filename, file_bytes=excel_buffer.getvalue())

    def generate_new_ssr_summary(self, sum_per_client_sheet: Dict) -> Dict:

        ssr_summary_dict = {}

        for client_name, values in sum_per_client_sheet.items():
            soh_po_sum = values.get("soh_value_sum") + values.get("po_cost_sum")

            ssr_summary_dict[client_name] = {
                "SOH VALUE": values.get("soh_value_sum"),
                "PO COST": values.get("po_cost_sum"),
                "SOH + PO COST": soh_po_sum,
                "SO COST": values.get("so_cost_sum"),
                "LIABILITY": soh_po_sum - values.get("so_cost_sum")
            }

        return ssr_summary_dict
                                
    def get_start_end_fiscal_year(self):
        if self.australia_now.month >= 7:
            fiscal_year_start = self.australia_now.year
            fiscal_year_end = self.australia_now.year + 1
        else:
            fiscal_year_start = self.australia_now.year - 1
            fiscal_year_end = self.australia_now.year  

        return fiscal_year_start, fiscal_year_end
    
    def get_target_ssr_summary_table_column(self, ssr_summary_df: pd.DataFrame) -> int | None:
        row = ssr_summary_df.iloc[1]  # 2nd row (Python index 1)

        for col_idx, val in row.items():
            try:
                # Try parsing the value as a date
                parsed = parse(str(val), fuzzy=True)
                if parsed.replace(tzinfo=timezone.utc) == self.australia_now:
                    return col_idx
                elif parsed.replace(tzinfo=timezone.utc) >= self.australia_now:
                    return col_idx - 1
            except (ValueError, TypeError) as e:
                logging.warning(f"[StockStatusReportOps] Getting index for target date {self.australia_now} in SSR summary warning: {e}")
                continue

        return None

    @staticmethod
    def get_target_client_row(ssr_summary_df: pd.DataFrame, ssr_customer_label: str, start_idx: int = 0) -> int | None:
        idx_list = ssr_summary_df.iloc[start_idx:].index[ssr_summary_df.iloc[start_idx:, 0] == ssr_customer_label].tolist()

        if len(idx_list) > 0:
            return idx_list[0]

        return None

    def get_ssr_summary_and_product_list(self) -> Tuple[bytes, pd.DataFrame]:
        logging.info("[StockStatusReportOps] Retrieving SSR Summary and Product List")

        with ThreadPoolExecutor() as executor:
            ssr_summary_bytes_result = executor.submit(self.get_ssr_summary_bytes)
            product_list_result = executor.submit(self.get_product_list)

            ssr_summary_bytes = ssr_summary_bytes_result.result()
            product_list = product_list_result.result()

        return ssr_summary_bytes, product_list